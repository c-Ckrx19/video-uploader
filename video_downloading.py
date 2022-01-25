from openpyxl import load_workbook
import requests
import os
import threading


class VideoDownload:
    def __init__(self, video_location, parent_dir):
        self.video_location = video_location[0]
        self.parent_dir = parent_dir
        self.video_name_list = []

    def download(self, url):
        video_name = url.split('/')[-1]
        self.video_name_list.append(video_name)
        video_name = os.path.join(self.parent_dir, video_name)
        r = requests.get(url, stream=True)
        with open(video_name, 'wb') as f:
            for chunk in r.iter_content(chunk_size=8 * 1024):
                f.write(chunk)

    def run(self):
        video_loc_wb = load_workbook(self.video_location)
        video_loc_sheet = video_loc_wb.active
        url_list = [video_loc_sheet['A'+str(row_num)].value
                    for row_num in range(1, video_loc_sheet.max_row+1)]
        i = 0
        while i < len(url_list):
            p1 = threading.Thread(target=self.download, args=(url_list[i],))
            try:
                p2 = threading.Thread(target=self.download, args=(url_list[i+1],))
            except IndexError:
                p2 = threading.Thread()
            p1.start()
            p2.start()
            p1.join()
            p2.join()
            i += 2
        return self.video_name_list
