import os
from moviepy.video.io.ffmpeg_tools import ffmpeg_extract_subclip
from openpyxl import load_workbook


def time_conversion(text_time):
    # Convert time format from xx:xx:xx to seconds
    # e.g. 1:30:30 --> 5430
    time_list = [int(t) for t in text_time.split(":")]
    to_sec = 3600 * time_list[0] + 60 * time_list[1] + time_list[2]
    return to_sec


class VideoSplitter:
    def __init__(self, video_filename, stat_filename, part=''):
        self.video_filename = video_filename
        self.stat_filename = stat_filename
        self.part = part

    def split(self):
        stat_wb = load_workbook(self.stat_filename)
        stat_sheet = None
        if self.part == '':
            stat_sheet = stat_wb.active
        else:
            for sheet_name in stat_wb.sheetnames:
                if self.part in sheet_name:
                    stat_sheet = stat_wb[sheet_name]
                    break
        start_col = 'D'
        end_col = 'E'
        for row_num in range(2, stat_sheet.max_row+1):
            start_second = time_conversion(stat_sheet[start_col+str(row_num)].value)
            end_second = time_conversion(stat_sheet[end_col+str(row_num)].value)
            ffmpeg_extract_subclip(self.video_filename, start_second, 
                                   end_second, targetname=str(row_num - 1)+".mp4")
